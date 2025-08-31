import sqlite3
import pandas as pd
from openpyxl import load_workbook
import os
import shutil
from datetime import datetime
from utils.logger import logger

def detectar_colunas(df):
    """
    Detecta automaticamente as colunas relevantes no DataFrame
    Retorna um dicionário com os mapeamentos encontrados
    """
    mapeamento_colunas = {
        'numero': None,
        'nome': None,
        'localizacao': None,
        'situacao': None
    }
    
    # Mapeamento de possíveis nomes para cada coluna (em minúsculas)
    possiveis_nomes = {
        'numero': [
            'número do bem', 'numero do bem', 'nº do bem', 'n° do bem',
            'patrimonio', 'patrimônio', 'numero', 'número', 'código', 
            'codigo', 'id', 'número patrimonial', 'numero patrimonial',
            'número do patrimônio', 'numero do patrimonio', 'asset number',
            'patrimônio', 'código do bem', 'codigo do bem'
        ],
        'nome': [
            'nome', 'descrição', 'descricao', 'item', 'equipamento', 
            'bem', 'denominação', 'denominacao', 'designação', 'designacao',
            'especificação', 'especificacao', 'produto', 'material',
            'nome do item', 'nome do equipamento', 'nome do bem',
            'description', 'item name', 'equipment name'
        ],
        'localizacao': [
            'localização', 'localizacao', 'local', 'setor', 'departamento',
            'área', 'area', 'sala', 'ambiente', 'prédio', 'predio', 'bloco',
            'unidade', 'centro de custo', 'departamento', 'division',
            'location', 'setor', 'department', 'area'
        ],
        'situacao': [
            'situação', 'situacao', 'status', 'estado', 'condição',
            'condicao', 'estado de conservação', 'estado de conservacao',
            'status do bem', 'situação do bem', 'situacao do bem',
            'status', 'state', 'condition', 'situation'
        ]
    }
    
    # Converter nomes das colunas para minúsculas e remover espaços extras
    colunas_df = [str(col).strip().lower() for col in df.columns]
    
    logger.info(f"Colunas encontradas no Excel: {colunas_df}")
    
    # Procurar correspondências para cada coluna
    for coluna_alvo, possibilidades in possiveis_nomes.items():
        for possibilidade in possibilidades:
            # Verificar correspondência exata ou parcial
            for coluna_df in colunas_df:
                if possibilidade == coluna_df or possibilidade in coluna_df:
                    indice = colunas_df.index(coluna_df)
                    mapeamento_colunas[coluna_alvo] = df.columns[indice]
                    logger.info(f"Coluna '{coluna_alvo}' detectada como: '{df.columns[indice]}'")
                    break
            if mapeamento_colunas[coluna_alvo]:
                break
    
    return mapeamento_colunas

def normalizar_valor(valor):
    """
    Normaliza valores para evitar problemas de tipo e formato
    """
    if pd.isna(valor) or valor is None:
        return None
    
    # Converter para string e remover espaços extras
    valor_str = str(valor).strip()
    
    # Se for string vazia, retornar None
    if not valor_str:
        return None
    
    return valor_str

def importar_excel_para_sqlite(arquivo_excel, aba_nome='Estoque', caminho_sqlite=None, criar_backup=True):
    """
    Importa dados de um arquivo Excel para o banco SQLite com detecção automática de colunas
    """
    if caminho_sqlite is None:
        caminho_sqlite = "relatorios/controle_patrimonial.db"
    
    # Criar pasta se não existir
    os.makedirs(os.path.dirname(caminho_sqlite), exist_ok=True)
    
    try:
        # Fazer backup se solicitado e se o banco existir
        if criar_backup and os.path.exists(caminho_sqlite):
            backup_path = f"relatorios/backup_controle_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
            shutil.copy2(caminho_sqlite, backup_path)
            logger.info(f"Backup criado: {backup_path}")
            mensagem_backup = f"📦 Backup criado: {os.path.basename(backup_path)}"
        else:
            mensagem_backup = ""
        
        # Ler o arquivo Excel
        logger.info(f"Iniciando importação do arquivo: {arquivo_excel}")
        
        # Verificar se o arquivo existe
        if not os.path.exists(arquivo_excel):
            raise FileNotFoundError(f"Arquivo Excel não encontrado: {arquivo_excel}")
        
        # Verificar abas disponíveis
        wb = load_workbook(arquivo_excel, read_only=True)
        if aba_nome not in wb.sheetnames:
            abas_disponiveis = ", ".join(wb.sheetnames)
            raise ValueError(f"Aba '{aba_nome}' não encontrada. Abas disponíveis: {abas_disponiveis}")
        
        # Ler dados do Excel usando pandas
        df = pd.read_excel(arquivo_excel, sheet_name=aba_nome)
        
        # Verificar se há dados
        if df.empty:
            raise ValueError("O arquivo Excel está vazio ou não contém dados")
        
        # Detectar colunas automaticamente
        mapeamento = detectar_colunas(df)
        logger.info(f"Mapeamento de colunas detectado: {mapeamento}")
        
        # Verificar colunas obrigatórias
        if not mapeamento['numero']:
            colunas_disponiveis = list(df.columns)
            raise ValueError(f"""
            Não foi possível detectar a coluna do número do bem.
            Colunas disponíveis: {colunas_disponiveis}
            Nomes esperados: Número do Bem, Patrimonio, Código, etc.
            """)
        
        if not mapeamento['nome']:
            colunas_disponiveis = list(df.columns)
            raise ValueError(f"""
            Não foi possível detectar a coluna do nome.
            Colunas disponíveis: {colunas_disponiveis}
            Nomes esperados: Nome, Descrição, Item, etc.
            """)
        
        # Conectar ao SQLite
        conn = sqlite3.connect(caminho_sqlite)
        cursor = conn.cursor()
        
        # Criar tabela se não existir
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS bens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numero TEXT NOT NULL UNIQUE,
                nome TEXT NOT NULL,
                localizacao TEXT DEFAULT '',
                situacao TEXT DEFAULT 'Pendente',
                data_criacao DATETIME DEFAULT CURRENT_TIMESTAMP,
                data_localizacao DATETIME
            )
        """)
        
        # Limpar tabela existente
        cursor.execute("DELETE FROM bens")
        logger.info("Tabela limpa para nova importação")
        
        # Inserir dados
        registros_inseridos = 0
        registros_erro = 0
        registros_ignorados = 0
        
        for index, row in df.iterrows():
            try:
                # Pular linhas completamente vazias
                if row.isnull().all():
                    registros_ignorados += 1
                    continue
                
                # Obter valores
                numero_bem = normalizar_valor(row[mapeamento['numero']])
                nome = normalizar_valor(row[mapeamento['nome']])
                
                # Pular linhas com dados obrigatórios faltantes
                if not numero_bem or not nome:
                    registros_ignorados += 1
                    continue
                
                # Obter valores opcionais
                localizacao = ''
                if mapeamento['localizacao']:
                    localizacao_valor = normalizar_valor(row[mapeamento['localizacao']])
                    localizacao = localizacao_valor if localizacao_valor else ''
                
                situacao = 'Pendente'
                if mapeamento['situacao']:
                    situacao_valor = normalizar_valor(row[mapeamento['situacao']])
                    if situacao_valor:
                        # Tentar detectar automaticamente se está localizado
                        situacao_lower = situacao_valor.lower()
                        if any(termo in situacao_lower for termo in ['ok', 'localizado', 'encontrado', 'sim', 'yes', 'concluído']):
                            situacao = 'OK'
                        else:
                            situacao = situacao_valor
                
                # Inserir no banco
                cursor.execute("""
                    INSERT OR REPLACE INTO bens (numero, nome, localizacao, situacao)
                    VALUES (?, ?, ?, ?)
                """, (numero_bem, nome, localizacao, situacao))
                
                registros_inseridos += 1
                
                # Log a cada 100 registros
                if registros_inseridos % 100 == 0:
                    logger.info(f"Registros processados: {registros_inseridos}")
                
            except Exception as e:
                registros_erro += 1
                logger.warning(f"Erro na linha {index + 2}: {str(e)}")
                continue
        
        conn.commit()
        conn.close()
        
        # Mensagem de sucesso detalhada
        mensagem = f"✅ Importação concluída com sucesso!"
        mensagem += f"\n• 📊 Registros inseridos: {registros_inseridos}"
        
        if registros_erro > 0:
            mensagem += f"\n• ⚠️  Registros com erro: {registros_erro}"
        
        if registros_ignorados > 0:
            mensagem += f"\n• 🔄 Registros ignorados (vazios): {registros_ignorados}"
        
        if mensagem_backup:
            mensagem += f"\n• {mensagem_backup}"
        
        logger.info(mensagem)
        return True, mensagem
        
    except Exception as e:
        error_msg = f"❌ Erro na importação: {str(e)}"
        logger.error(error_msg)
        return False, error_msg

def verificar_estrutura_excel(arquivo_excel, aba_nome='Estoque'):
    """
    Verifica a estrutura do arquivo Excel antes da importação
    """
    try:
        wb = load_workbook(arquivo_excel, read_only=True)
        
        if aba_nome not in wb.sheetnames:
            return False, f"Aba '{aba_nome}' não encontrada"
        
        # Ler algumas linhas para verificar estrutura
        df = pd.read_excel(arquivo_excel, sheet_name=aba_nome, nrows=10)
        
        if df.empty:
            return False, "O arquivo Excel está vazio"
        
        # Detectar colunas automaticamente
        mapeamento = detectar_colunas(df)
        
        # Verificar colunas mínimas
        if not mapeamento['numero']:
            return False, "Coluna do número do bem não encontrada"
        
        if not mapeamento['nome']:
            return False, "Coluna do nome não encontrada"
        
        # Preparar mensagem detalhada
        colunas_detectadas = []
        for chave, valor in mapeamento.items():
            if valor:
                colunas_detectadas.append(f"{chave}: '{valor}'")
        
        return True, f"Estrutura válida. Colunas detectadas: {', '.join(colunas_detectadas)}"
        
    except Exception as e:
        return False, f"Erro na verificação: {str(e)}"

def obter_colunas_excel(arquivo_excel, aba_nome='Estoque'):
    """
    Retorna as colunas disponíveis no arquivo Excel
    """
    try:
        df = pd.read_excel(arquivo_excel, sheet_name=aba_nome, nrows=1)
        colunas = list(df.columns)
        
        # Adicionar informações de tipo
        df_sample = pd.read_excel(arquivo_excel, sheet_name=aba_nome, nrows=5)
        info_colunas = []
        
        for coluna in colunas:
            tipo = str(df_sample[coluna].dtype)
            valores_exemplo = df_sample[coluna].head(3).tolist()
            info_colunas.append(f"'{coluna}' (Tipo: {tipo}, Exemplo: {valores_exemplo})")
        
        return info_colunas
        
    except Exception as e:
        logger.error(f"Erro ao obter colunas: {str(e)}")
        return [f"Erro: {str(e)}"]

def criar_template_excel(caminho_saida):
    """
    Cria um template de Excel com a estrutura esperada
    """
    try:
        # Dados de exemplo
        dados = [
            ['Computador Dell', '1001', 'Sala 101', 'OK'],
            ['Monitor LG', '1002', 'Sala 102', 'Pendente'],
            ['Impressora HP', '1003', 'Recepção', 'OK'],
            ['Telefone IP', '1004', 'Escritório', 'Pendente']
        ]
        
        # Criar DataFrame
        df = pd.DataFrame(dados, columns=['Nome', 'Número do Bem', 'Localização', 'Situação'])
        
        # Salvar como Excel
        df.to_excel(caminho_saida, index=False)
        
        return True, f"Template criado: {caminho_saida}"
        
    except Exception as e:
        return False, f"Erro ao criar template: {str(e)}"